from openpyxl import Workbook
from openpyxl.styles import Font
from typing import Dict, Any
import re
import os
import ast 
#from rules_loader.py import load_validation_rules

METADATA_KEYS = [
    "description", "verifies", "verifies_tds", "id", "domain",
    "testType", "derivationTechnique", "coveredArchitectureElements",
    "ASIL", "status", "priority"
]

EXTENDED_METADATA_KEYS = [
    "testingSupplier", "feature", "max_severity", "zedis"
]
def extract_info(file_path: str) -> Dict[str, Any]:
    with open(file_path, "r", encoding="utf-8") as f:
        source = f.read()
    tree = ast.parse(source)

    result = {
        "file_name": os.path.basename(file_path),
        "class_name": None,
    }

    # Pré-remplir les clés avec None(initialisation des mes variables )
    for key in METADATA_KEYS:
        result[f"metadata:{key}"] = None
    for key in EXTENDED_METADATA_KEYS:
        result[f"extended_metadata:{key}"] = None

    for node in ast.walk(tree):
        if isinstance(node, ast.ClassDef):
            if not re.match(r"^Test[A-Z][a-zA-Z0-9]*$", node.name):
                print(f" Fichier '{os.path.basename(file_path)}': Classe '{node.name}' ne respecte pas le PascalCase avec préfixe 'Test'")

            result["class_name"] = node.name
            for decorator in node.decorator_list:
                if isinstance(decorator, ast.Call) and isinstance(decorator.func, ast.Attribute):
                    prefix = ""
                    if decorator.func.attr == "metadata":
                        prefix = "metadata"
                    elif decorator.func.attr == "extended_metadata":
                        prefix = "extended_metadata"
                    else:
                        continue

                    for kw in decorator.keywords:
                        key = kw.arg
                        try:
                            value = ast.literal_eval(kw.value)#retourne a python expression
                        except Exception:
                            value = None
                        result[f"{prefix}:{key}"] = value
    return result

def write_results_to_excel(results: Dict[int, Dict[str, Any]], output_file: str, validation_rules: Dict[str, Dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Extraction Metadata"

    headers = ["file_name", "class_name"]
    headers += [f"metadata:{k}" for k in METADATA_KEYS]
    headers += [f"extended_metadata:{k}" for k in EXTENDED_METADATA_KEYS]

    ws.append(headers)
    for _, info in results.items():
        row = [info.get("file_name", ""), info.get("class_name", "")]
        for key in headers[2:]:
            val = info.get(key, "")
            if isinstance(val, (list, dict)):
                val = str(val)
            elif isinstance(val, bool):
                val = "true" if val else "false"
            row.append(val if val is not None else "")
        ws.append(row)

    validate_and_colorize_excel(ws, validation_rules)
    wb.save(output_file)
    print(f"\n{len(results)} fichiers analysés")
    print(f"\nRésultats exportés dans : {output_file}")


def load_validation_rules(rules_file: str) -> Dict[str, Dict[str, Any]]:
    rules = {}
    with open(rules_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()#efface les espace w les retoure a la ligne zeydin 
            if not line or line.startswith("#"):
                continue

            parts = line.split(";")#la fonction split(";") découpe la ligne en plusieurs morceaux.->[a:b,c:d]ect...
            rule = {}
            for part in parts[1:]:  # ignorer le nom du champ pour l’instant
                key, value = part.split(":", 1)
                key = key.strip().lower()
                value = value.strip()
                if key == "mandatory":
                    rule["mandatory"] = value.lower() == "yes"
                elif key == "type":
                    rule["type"] = value.lower()
                elif key == "possible values":
                    if value.lower() == "any":
                        rule["values"] = None
                    else:
                        vals = value.split(",")
                        # Convertir automatiquement les types connus
                        if rule.get("type") == "integer":
                            rule["values"] = [int(v) for v in vals]
                        elif rule.get("type") == "boolean":
                            #rule["values"] = [v.lower() == "true" for v in vals]
                            rule["values"] = [str(v).lower() for v in vals]
                        else:
                            rule["values"] = [v.strip() for v in vals]
            rules[parts[0].strip().lower()] = rule
    return rules




def validate_and_colorize_excel(ws, rules: Dict[str, Dict[str, Any]]) -> None:
    red_font = Font(color="FF0000")
    green_font = Font(color="008000")
    sky_blue_font = Font(color="87CEEB")
    brown_font = Font(color="A52A2A")

    headers = [str(cell.value).lower() for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):#hna 9e3ed ya3mel f iteration mte3 les columns mais fonction esemha iter_rows dans "openpyxl"
        file_name, class_name = row[0], row[1]
        if not re.match(r"^[a-z_][a-z0-9_]*\.py$", str(file_name.value)):
            file_name.font = sky_blue_font
        else:
            file_name.font = green_font
        if not re.match(r"^Test[A-Z][a-zA-Z0-9]*$", str(class_name.value)):
            class_name.font = brown_font
        else:
            class_name.font = green_font

        for idx, cell in enumerate(row):
            header = headers[idx]
            rule = rules.get(header)
            if not rule:
                continue
            value = cell.value
            if value in [None, ""] and rule.get("mandatory"):
                cell.font = red_font
                continue

            expected_type = rule.get("type")
            valid_type = True
            try:
                if expected_type == "string":
                    valid_type = isinstance(value, str)
                elif expected_type == "integer":
                    int(value)
                elif expected_type == "boolean":
                    valid_type = str(value).lower() in ["true", "false"]
                elif expected_type == "list":
                    val_list = eval(value) if isinstance(value, str) else value
                    valid_type = isinstance(val_list, list)
            except:
                valid_type = False

            if not valid_type:
                cell.font = red_font
                continue

            allowed_values = rule.get("values")
            if allowed_values:
                if expected_type == "list":
                    val_list = eval(value) if isinstance(value, str) else value
                    if not all(str(v).lower() in map(str.lower, map(str, allowed_values)) for v in val_list):
                        cell.font = red_font
                    else:
                        cell.font = green_font
                elif str(value).lower() not in map(str.lower, map(str, allowed_values)):
                    cell.font = red_font
                else:
                    cell.font = green_font
            else:
                cell.font = green_font



if __name__ == "__main__":
    import sys
    target = sys.argv[1]
    output_file = sys.argv[2]
    rules_file = sys.argv[3]
    VALIDATION_RULES = load_validation_rules(rules_file)
    paths = []
    if os.path.isdir(target):
        for root, _, files in os.walk(target):
            for f in files:
                if f.endswith(".py") and f.startswith("test"):
                    paths.append(os.path.join(root, f))
    elif target.endswith('.py') and os.path.basename(target).startswith("test"):
        paths.append(target)
    else:
        print("Aucun fichier .py commençant par 'test' trouvé.")
        sys.exit(1)
    

    results = {}
    for i, path in enumerate(paths, start=1):
        results[i] = extract_info(path)

    write_results_to_excel(results, output_file, VALIDATION_RULES)